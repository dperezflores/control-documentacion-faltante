from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from services.excel_service import (
    FALTANTES_SHEET,
    _headers,
    _normalize_single_line_text,
    _rebuild_visible_cell,
    _split_visible_documents,
)


DEFAULT_INPUT = Path("data/Documentacion_faltante.xlsx")
DEFAULT_OUTPUT = Path("Documentacion_faltante_corregido.xlsx")


def _raw_multiline_parts(value) -> list[str]:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    return [line.strip() for line in text.split("\n") if line.strip()]


def repair_multiline_documents(input_path: Path, output_path: Path) -> int:
    wb = load_workbook(input_path)

    if FALTANTES_SHEET not in wb.sheetnames:
        print(f"No existe la hoja técnica {FALTANTES_SHEET}. No hay nada que corregir.")
        wb.save(output_path)
        print(f"Se generó una copia sin cambios en: {output_path}")
        return 0

    tech = wb[FALTANTES_SHEET]
    headers = _headers(tech)

    affected_contracts: set[tuple[str, str]] = set()
    ghost_parts_by_contract: dict[tuple[str, str], set[str]] = {}
    normalized_rows: list[dict] = []

    for row in range(2, tech.max_row + 1):
        requirement = str(tech.cell(row, headers["Requerimiento"]).value or "").strip()
        contract = str(tech.cell(row, headers["Contrato"]).value or "").strip()
        document = tech.cell(row, headers["Documento"]).value
        specification = (
            tech.cell(row, headers["Especificacion"]).value
            if "Especificacion" in headers
            else ""
        )

        new_document = _normalize_single_line_text(document)
        new_specification = _normalize_single_line_text(specification)

        changed_fields = []
        if str(document or "") != new_document:
            changed_fields.append("Documento")
        if str(specification or "") != new_specification:
            changed_fields.append("Especificacion")

        if not changed_fields:
            continue

        key = (requirement, contract)
        affected_contracts.add(key)
        ghost_parts = ghost_parts_by_contract.setdefault(key, set())

        if "Documento" in changed_fields:
            ghost_parts.update(_raw_multiline_parts(document))
            tech.cell(row, headers["Documento"]).value = new_document

        if "Especificacion" in changed_fields:
            ghost_parts.update(_raw_multiline_parts(specification))
            tech.cell(row, headers["Especificacion"]).value = new_specification

        normalized_rows.append(
            {
                "fila": row,
                "requerimiento": requirement,
                "contrato": contract,
                "campos": ", ".join(changed_fields),
                "documento_antes": str(document or ""),
                "documento_despues": new_document,
                "especificacion_antes": str(specification or ""),
                "especificacion_despues": new_specification,
            }
        )

    if not normalized_rows:
        print("No se encontraron filas técnicas con saltos de línea.")
        wb.save(output_path)
        print(f"Se generó una copia sin cambios en: {output_path}")
        return 0

    print("REPORTE DE NORMALIZACIÓN")
    print("=" * 80)
    for item in normalized_rows:
        print(
            f"Fila técnica {item['fila']} | "
            f"Requerimiento: {item['requerimiento']} | "
            f"Contrato: {item['contrato']} | "
            f"Campos: {item['campos']}"
        )
        if "Documento" in item["campos"]:
            print(f"  Documento antes : {item['documento_antes']!r}")
            print(f"  Documento después: {item['documento_despues']!r}")
        if "Especificacion" in item["campos"]:
            print(f"  Especificación antes : {item['especificacion_antes']!r}")
            print(f"  Especificación después: {item['especificacion_despues']!r}")
        print()

    print("CONTRATOS A RECONSTRUIR")
    print("=" * 80)
    for requirement, contract in sorted(affected_contracts):
        print(f"- {requirement} | {contract}")

    print()
    print("El archivo original NO será sobrescrito.")
    print(f"Archivo de salida: {output_path}")
    print()

    # Antes de usar _rebuild_visible_cell(), retiramos de la celda visible
    # únicamente los fragmentos generados por los saltos de línea de los
    # registros técnicos afectados. Esto evita conservar como 'manuales'
    # las líneas fantasma y deja que la lógica normal reconstruya la celda.
    for requirement, contract in affected_contracts:
        if requirement not in wb.sheetnames:
            continue

        visible_ws = wb[requirement]
        target_row = None
        for row in range(8, visible_ws.max_row + 1):
            if str(visible_ws.cell(row, 2).value or "").strip() == contract:
                target_row = row
                break

        if target_row is None:
            continue

        current_lines = _split_visible_documents(visible_ws.cell(target_row, 5).value)
        ghost_parts = ghost_parts_by_contract.get((requirement, contract), set())
        filtered_lines = [line for line in current_lines if line not in ghost_parts]
        visible_ws.cell(target_row, 5).value = "\n".join(filtered_lines)

        _rebuild_visible_cell(wb, requirement, contract)

    wb.save(output_path)
    print("Corrección terminada.")
    print(f"Filas técnicas normalizadas: {len(normalized_rows)}")
    print(f"Contratos reconstruidos: {len(affected_contracts)}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Normaliza textos multilínea en _APP_FALTANTES y reconstruye "
            "las celdas visibles afectadas sin sobrescribir el archivo original."
        )
    )
    parser.add_argument(
        "archivo",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT,
        help=(
            "Excel operativo descargado/exportado desde Google Drive. "
            "Por defecto: data/Documentacion_faltante.xlsx"
        ),
    )
    parser.add_argument(
        "--salida",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Ruta del nuevo Excel corregido.",
    )
    args = parser.parse_args()

    if not args.archivo.exists():
        print(
            "ERROR: no existe el archivo de entrada. Descarga/exporta primero "
            "el Excel operativo real desde Google Drive y vuelve a ejecutar."
        )
        print(f"Ruta esperada: {args.archivo}")
        return 2

    return repair_multiline_documents(args.archivo, args.salida)


if __name__ == "__main__":
    raise SystemExit(main())
