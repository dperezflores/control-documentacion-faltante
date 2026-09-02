from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_FILE = Path("data/Documentacion_faltante.xlsx")


def split_visible_documents_raw(value) -> list[str]:
    """Replica la separación visible SIN deduplicar para fines de auditoría."""
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    return [
        line.strip(" •\t")
        for line in text.split("\n")
        if line.strip(" •\t")
    ]


def audit_duplicates(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=False)
    findings: list[dict] = []

    for requirement in wb.sheetnames:
        if not requirement.startswith("Req_"):
            continue

        ws = wb[requirement]
        for row in range(8, ws.max_row + 1):
            contract = str(ws.cell(row, 2).value or "").strip()
            if not contract:
                continue

            documents = split_visible_documents_raw(ws.cell(row, 5).value)
            counts = Counter(documents)
            for text, count in counts.items():
                if count > 1:
                    findings.append(
                        {
                            "requerimiento": requirement,
                            "contrato": contract,
                            "texto_duplicado": text,
                            "apariciones": count,
                            "fila_excel": row,
                        }
                    )

    return findings


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Audita líneas duplicadas en la columna Documentación faltante "
            "de las hojas Req_* sin modificar el Excel."
        )
    )
    parser.add_argument(
        "archivo",
        nargs="?",
        type=Path,
        default=DEFAULT_FILE,
        help="Ruta del Excel operativo.",
    )
    args = parser.parse_args()

    if not args.archivo.exists():
        print(f"ERROR: no existe el archivo: {args.archivo}")
        return 2

    findings = audit_duplicates(args.archivo)

    if not findings:
        print("No se encontraron líneas duplicadas dentro de un mismo contrato.")
        return 0

    print(f"Se encontraron {len(findings)} duplicado(s):")
    print()
    for item in findings:
        print(
            f"- Requerimiento: {item['requerimiento']} | "
            f"Contrato: {item['contrato']} | "
            f"Fila: {item['fila_excel']} | "
            f"Apariciones: {item['apariciones']}"
        )
        print(f"  Texto duplicado: {item['texto_duplicado']}")
        print()

    print("El archivo NO fue modificado.")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
