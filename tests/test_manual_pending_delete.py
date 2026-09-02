from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

from services.excel_service import (
    _manual_id,
    delete_pending_records,
    manual_visible_documents,
)


def _build_manual_only_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Req_TEST"

    # La lógica productiva empieza a leer contratos en la fila 8.
    ws.cell(8, 1).value = 1
    ws.cell(8, 2).value = "CT-001"
    ws.cell(8, 3).value = "Obra de prueba"
    ws.cell(8, 4).value = "Contratista"
    ws.cell(8, 5).value = "Documento manual genuino"
    ws.cell(8, 6).value = "AUD"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_delete_genuine_manual_pending_does_not_regenerate():
    data = _build_manual_only_workbook()
    record_id = _manual_id("Req_TEST", "CT-001", "Documento manual genuino")

    updated = delete_pending_records(data, "Req_TEST", [record_id])

    wb = load_workbook(BytesIO(updated))
    visible_value = wb["Req_TEST"].cell(8, 5).value
    assert visible_value in (None, "")

    # Una lectura posterior tampoco debe reconstruir/regenerar el manual.
    assert manual_visible_documents(updated, "Req_TEST", "CT-001") == []
