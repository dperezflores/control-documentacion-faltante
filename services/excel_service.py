from __future__ import annotations

from datetime import datetime
from hashlib import sha1
from io import BytesIO
from typing import Iterable
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

FALTANTES_SHEET = "_APP_FALTANTES"
CORTES_SHEET = "_APP_CORTES"
OFICIOS_SHEET = "_APP_OFICIOS"

FALTANTES_HEADERS = [
    "ID", "Requerimiento", "Contrato", "Procedimiento", "Codigo_documento",
    "Documento", "Especificacion", "Fecha_deteccion", "Auditor", "Corte", "Origen"
]
CORTES_HEADERS = ["ID", "Requerimiento", "Corte", "Fecha", "Creado_por", "Documentos"]
OFICIOS_HEADERS = ["ID", "Requerimiento", "Fecha", "Cortes", "Referencia", "Creado_por"]
LOCAL_TZ = ZoneInfo("America/Mexico_City")


def configure_timezone(timezone_name: str) -> None:
    global LOCAL_TZ
    LOCAL_TZ = ZoneInfo(timezone_name)


def _load(data: bytes):
    return load_workbook(BytesIO(data))


def _save(wb) -> bytes:
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _ensure_headers(ws, headers: list[str]) -> None:
    current = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not any(current):
        ws.append(headers)
        return
    for header in headers:
        if header not in current:
            ws.cell(1, ws.max_column + 1).value = header
            current.append(header)


def _headers(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def _ensure_technical_sheets(wb) -> None:
    specs = [
        (FALTANTES_SHEET, FALTANTES_HEADERS),
        (CORTES_SHEET, CORTES_HEADERS),
        (OFICIOS_SHEET, OFICIOS_HEADERS),
    ]
    for name, headers in specs:
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(headers)
        else:
            ws = wb[name]
            _ensure_headers(ws, headers)
        ws.sheet_state = "hidden"


def _effective_document(documento, especificacion) -> str:
    doc = str(documento or "").strip()
    spec = str(especificacion or "").strip()
    if doc and spec:
        return f"{doc} ({spec})"
    return doc or spec

def _record_origin(row: dict) -> str:
    explicit = str(row.get("Origen") or "").strip()
    if explicit:
        return explicit
    if str(row.get("Procedimiento") or "").strip().upper() == "MANUAL":
        return "Registro previo / manual"
    return "Aplicación"


def _split_visible_documents(value) -> list[str]:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = [
        line.strip(" •\t")
        for line in text.split("\n")
        if line.strip(" •\t")
    ]
    return list(dict.fromkeys(lines))


def _manual_id(requirement: str, contract: str, text: str) -> str:
    digest = sha1(f"{requirement}|{contract}|{text}".encode("utf-8")).hexdigest()[:16]
    return f"MANUAL::{digest}"


def _manual_pending_rows_from_wb(wb, requirement: str) -> list[dict]:
    _ensure_technical_sheets(wb)
    visible_ws = wb[requirement]
    tech = wb[FALTANTES_SHEET]
    h = _headers(tech)

    technical_by_contract: dict[str, set[str]] = {}
    for r in range(2, tech.max_row + 1):
        if tech.cell(r, h["Requerimiento"]).value != requirement:
            continue
        contract = str(tech.cell(r, h["Contrato"]).value or "").strip()
        technical_by_contract.setdefault(contract, set()).add(
            _effective_document(
                tech.cell(r, h["Documento"]).value,
                tech.cell(r, h["Especificacion"]).value if "Especificacion" in h else "",
            )
        )

    result = []
    for r in range(8, visible_ws.max_row + 1):
        contract = str(visible_ws.cell(r, 2).value or "").strip()
        if not contract:
            continue
        auditor = str(visible_ws.cell(r, 6).value or "").strip()
        visible_docs = _split_visible_documents(visible_ws.cell(r, 5).value or "")
        technical_docs = technical_by_contract.get(contract, set())
        for text in visible_docs:
            if text in technical_docs:
                continue
            result.append({
                "id": _manual_id(requirement, contract, text),
                "requerimiento": requirement,
                "contrato": contract,
                "codigo": "",
                "documento": text,
                "especificacion": "",
                "solicitud": text,
                "fecha": None,
                "auditor": auditor,
                "origen": "Registro previo / manual",
                "_manual_text": text,
            })
    return result


def manual_visible_documents(data: bytes, requirement: str, contract: str) -> list[str]:
    wb = _load(data)
    return [
        row["solicitud"]
        for row in _manual_pending_rows_from_wb(wb, requirement)
        if row["contrato"] == contract
    ]


def _technical_docs_for_contract(wb, requirement: str, contract: str) -> list[str]:
    tech = wb[FALTANTES_SHEET]
    h = _headers(tech)
    docs: list[str] = []
    for r in range(2, tech.max_row + 1):
        if (
            tech.cell(r, h["Requerimiento"]).value == requirement
            and str(tech.cell(r, h["Contrato"]).value or "").strip() == contract
        ):
            text = _effective_document(
                tech.cell(r, h["Documento"]).value,
                tech.cell(r, h["Especificacion"]).value if "Especificacion" in h else "",
            )
            if text and text not in docs:
                docs.append(text)
    return docs


def _write_visible_cell(wb, requirement: str, contract: str, manual_docs: list[str]) -> None:
    visible_ws = wb[requirement]
    target_row = None
    for r in range(8, visible_ws.max_row + 1):
        if str(visible_ws.cell(r, 2).value or "").strip() == contract:
            target_row = r
            break
    if target_row is None:
        raise ValueError("No se encontró el contrato en la hoja seleccionada.")

    technical_docs = _technical_docs_for_contract(wb, requirement, contract)
    combined = list(dict.fromkeys([*manual_docs, *technical_docs]))
    visible_ws.cell(target_row, 5).value = "\n".join(combined)


def _rebuild_visible_cell(wb, requirement: str, contract: str) -> None:
    visible_ws = wb[requirement]
    current_visible = ""
    for r in range(8, visible_ws.max_row + 1):
        if str(visible_ws.cell(r, 2).value or "").strip() == contract:
            current_visible = visible_ws.cell(r, 5).value or ""
            break

    technical_docs = set(_technical_docs_for_contract(wb, requirement, contract))
    manual_docs = [x for x in _split_visible_documents(current_visible) if x not in technical_docs]
    _write_visible_cell(wb, requirement, contract, manual_docs)


def _append_manual_to_tech(wb, requirement: str, manual_row: dict, cut_value) -> None:
    tech = wb[FALTANTES_SHEET]
    h = _headers(tech)
    row = [None] * tech.max_column
    values = {
        "ID": uuid4().hex,
        "Requerimiento": requirement,
        "Contrato": manual_row["contrato"],
        "Procedimiento": "MANUAL",
        "Codigo_documento": f"MANUAL_{manual_row['id'].split('::')[-1]}",
        "Documento": manual_row["solicitud"],
        "Especificacion": "",
        "Fecha_deteccion": datetime.now(LOCAL_TZ).replace(tzinfo=None, microsecond=0),
        "Auditor": manual_row.get("auditor") or "",
        "Corte": cut_value,
        "Origen": "Registro previo / manual",
    }
    for name, value in values.items():
        row[h[name] - 1] = value
    tech.append(row)


def _update_cut_count(wb, requirement: str, cut_number: int) -> None:
    falt = wb[FALTANTES_SHEET]
    cuts = wb[CORTES_SHEET]
    fh = _headers(falt)
    ch = _headers(cuts)
    count = 0
    for r in range(2, falt.max_row + 1):
        if falt.cell(r, fh["Requerimiento"]).value == requirement:
            try:
                current = int(falt.cell(r, fh["Corte"]).value)
            except (TypeError, ValueError):
                continue
            if current == int(cut_number):
                count += 1
    for r in range(2, cuts.max_row + 1):
        if cuts.cell(r, ch["Requerimiento"]).value == requirement and int(cuts.cell(r, ch["Corte"]).value) == int(cut_number):
            cuts.cell(r, ch["Documentos"]).value = count
            break


def list_requirements(data: bytes) -> list[str]:
    wb = _load(data)
    return [s for s in wb.sheetnames if s.startswith("Req_")]


def list_contracts(data: bytes, requirement: str) -> list[dict]:
    wb = _load(data)
    ws = wb[requirement]
    rows = []
    for row in range(8, ws.max_row + 1):
        number = ws.cell(row, 1).value
        contract = ws.cell(row, 2).value
        if number is None or not contract:
            continue
        rows.append({
            "row": row,
            "numero": number,
            "contrato": str(contract).strip(),
            "obra": ws.cell(row, 3).value or "",
            "contratista": ws.cell(row, 4).value or "",
            "faltantes_visible": ws.cell(row, 5).value or "",
            "auditor": ws.cell(row, 6).value or "",
            "ulop": ws.cell(row, 7).value or "",
        })
    return rows


def load_catalog(data: bytes, procedure: str) -> list[dict]:
    wb = _load(data)
    ws = wb[procedure]
    items = []
    for row in range(3, ws.max_row + 1):
        code = ws.cell(row, 1).value
        concept = ws.cell(row, 2).value
        if not code or not concept:
            continue
        prefix = str(code).split("_")[0]
        items.append({"codigo": str(code).strip(), "documento": str(concept).strip(), "etapa": prefix})
    return items


def get_faltantes(data: bytes, requirement: str, contract: str | None = None) -> list[dict]:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    ws = wb[FALTANTES_SHEET]
    h = _headers(ws)
    result = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, h["Requerimiento"]).value != requirement:
            continue
        if contract is not None and str(ws.cell(r, h["Contrato"]).value or "").strip() != contract:
            continue
        row = {name: ws.cell(r, col).value for name, col in h.items()}
        row["_row"] = r
        row["Documento_efectivo"] = _effective_document(row.get("Documento"), row.get("Especificacion"))
        result.append(row)
    return result


def document_history(data: bytes, requirement: str, contract: str) -> dict[str, dict]:
    history: dict[str, dict] = {}
    for row in get_faltantes(data, requirement, contract):
        code = str(row.get("Codigo_documento") or "")
        if not code:
            continue
        entry = history.setdefault(
            code,
            {
                "cuts": [],
                "pending": False,
                "documento": row.get("Documento"),
                "origen": _record_origin(row),
            },
        )
        cut = row.get("Corte")
        cut_text = str(cut)
        if cut_text == "PENDIENTE":
            entry["pending"] = True
        elif cut_text.startswith("GENERAL:"):
            try:
                general_num = int(cut_text.split(":", 1)[1])
                label = f"General {general_num}"
                if label not in entry["cuts"]:
                    entry["cuts"].append(label)
            except (TypeError, ValueError):
                pass
        else:
            try:
                cut_num = int(cut)
                if cut_num not in entry["cuts"]:
                    entry["cuts"].append(cut_num)
            except (TypeError, ValueError):
                pass
    for entry in history.values():
        def _sort_key(value):
            if isinstance(value, int):
                return (0, value)
            text = str(value)
            if text.startswith("General "):
                try:
                    return (1, int(text.split(" ", 1)[1]))
                except ValueError:
                    return (1, text)
            return (2, text)
        entry["cuts"].sort(key=_sort_key)
    return history


def add_faltantes(data: bytes, requirement: str, contract: str, procedure: str, auditor: str, selected: Iterable[dict]) -> bytes:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    tech = wb[FALTANTES_SHEET]
    h = _headers(tech)

    pending_codes = set()
    for r in range(2, tech.max_row + 1):
        if tech.cell(r, h["Requerimiento"]).value == requirement and str(tech.cell(r, h["Contrato"]).value or "").strip() == contract:
            if str(tech.cell(r, h["Corte"]).value) == "PENDIENTE":
                pending_codes.add(str(tech.cell(r, h["Codigo_documento"]).value))

    now = datetime.now(LOCAL_TZ).replace(tzinfo=None, microsecond=0)
    for item in selected:
        if item["codigo"] in pending_codes:
            continue
        row = [None] * tech.max_column
        values = {
            "ID": uuid4().hex,
            "Requerimiento": requirement,
            "Contrato": contract,
            "Procedimiento": procedure,
            "Codigo_documento": item["codigo"],
            "Documento": item["documento"],
            "Especificacion": str(item.get("especificacion") or "").strip(),
            "Fecha_deteccion": now,
            "Auditor": auditor,
            "Corte": "PENDIENTE",
            "Origen": "Aplicación",
        }
        for name, value in values.items():
            row[h[name] - 1] = value
        tech.append(row)
        pending_codes.add(item["codigo"])

    _rebuild_visible_cell(wb, requirement, contract)
    return _save(wb)


def pending_summary(data: bytes, requirement: str) -> list[dict]:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    tech = wb[FALTANTES_SHEET]
    h = _headers(tech)
    result = []

    for r in range(2, tech.max_row + 1):
        if tech.cell(r, h["Requerimiento"]).value != requirement:
            continue
        if str(tech.cell(r, h["Corte"]).value) != "PENDIENTE":
            continue
        documento = tech.cell(r, h["Documento"]).value
        especificacion = tech.cell(r, h["Especificacion"]).value if "Especificacion" in h else ""
        result.append({
            "id": tech.cell(r, h["ID"]).value,
            "requerimiento": requirement,
            "contrato": tech.cell(r, h["Contrato"]).value,
            "codigo": tech.cell(r, h["Codigo_documento"]).value,
            "documento": documento,
            "especificacion": especificacion or "",
            "solicitud": _effective_document(documento, especificacion),
            "fecha": tech.cell(r, h["Fecha_deteccion"]).value,
            "auditor": tech.cell(r, h["Auditor"]).value,
            "origen": _record_origin({
                "Origen": tech.cell(r, h["Origen"]).value if "Origen" in h else "",
                "Procedimiento": tech.cell(r, h["Procedimiento"]).value,
            }),
        })

    result.extend(_manual_pending_rows_from_wb(wb, requirement))
    return result


def list_cuts(data: bytes, requirement: str) -> list[dict]:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    ws = wb[CORTES_SHEET]
    h = _headers(ws)
    result = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, h["Requerimiento"]).value == requirement:
            result.append({
                "id": ws.cell(r, h["ID"]).value,
                "corte": int(ws.cell(r, h["Corte"]).value),
                "fecha": ws.cell(r, h["Fecha"]).value,
                "creado_por": ws.cell(r, h["Creado_por"]).value,
                "documentos": int(ws.cell(r, h["Documentos"]).value or 0),
            })
    return sorted(result, key=lambda x: x["corte"])


def cut_details(data: bytes, requirement: str, cut_number: int) -> list[dict]:
    result = []
    for row in get_faltantes(data, requirement):
        try:
            current_cut = int(row.get("Corte"))
        except (TypeError, ValueError):
            continue
        if current_cut == int(cut_number):
            result.append({
                "id": row.get("ID"),
                "requerimiento": requirement,
                "contrato": row.get("Contrato"),
                "auditor": row.get("Auditor"),
                "codigo": row.get("Codigo_documento"),
                "documento": row.get("Documento"),
                "especificacion": row.get("Especificacion") or "",
                "solicitud": row.get("Documento_efectivo"),
                "origen": _record_origin(row),
                "fecha": row.get("Fecha_deteccion"),
                "corte": current_cut,
            })
    return result


def create_cut(data: bytes, requirement: str, user: str, cut_date) -> bytes:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    falt = wb[FALTANTES_SHEET]
    cuts = wb[CORTES_SHEET]
    fh = _headers(falt)
    ch = _headers(cuts)

    existing = [
        int(cuts.cell(r, ch["Corte"]).value)
        for r in range(2, cuts.max_row + 1)
        if cuts.cell(r, ch["Requerimiento"]).value == requirement and cuts.cell(r, ch["Corte"]).value
    ]
    next_cut = max(existing, default=0) + 1

    manual_rows = _manual_pending_rows_from_wb(wb, requirement)
    for manual_row in manual_rows:
        _append_manual_to_tech(wb, requirement, manual_row, next_cut)

    affected = len(manual_rows)
    for r in range(2, falt.max_row + 1):
        if (
            falt.cell(r, fh["Requerimiento"]).value == requirement
            and str(falt.cell(r, fh["Corte"]).value) == "PENDIENTE"
        ):
            falt.cell(r, fh["Corte"]).value = next_cut
            affected += 1

    if affected == 0:
        raise ValueError("No hay documentación pendiente para crear un corte.")

    row = [None] * cuts.max_column
    values = {
        "ID": uuid4().hex,
        "Requerimiento": requirement,
        "Corte": next_cut,
        "Fecha": cut_date,
        "Creado_por": user,
        "Documentos": affected,
    }
    for name, value in values.items():
        row[ch[name] - 1] = value
    cuts.append(row)
    return _save(wb)


def delete_pending_records(data: bytes, requirement: str, record_ids: Iterable[str]) -> bytes:
    ids = {str(x) for x in record_ids}
    if not ids:
        return data

    wb = _load(data)
    _ensure_technical_sheets(wb)
    ws = wb[FALTANTES_SHEET]
    h = _headers(ws)

    manual_rows = {row["id"]: row for row in _manual_pending_rows_from_wb(wb, requirement)}
    manual_by_contract: dict[str, list[str]] = {}
    for row in manual_rows.values():
        manual_by_contract.setdefault(row["contrato"], []).append(row["solicitud"])

    # Eliminar registros manuales seleccionados de la celda visible.
    for record_id in ids:
        manual = manual_rows.get(record_id)
        if manual:
            contract = manual["contrato"]
            manual_by_contract[contract] = [
                x for x in manual_by_contract.get(contract, [])
                if x != manual["solicitud"]
            ]

    affected_contracts: set[str] = set()
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        if (
            ws.cell(r, h["Requerimiento"]).value == requirement
            and str(ws.cell(r, h["ID"]).value) in ids
            and str(ws.cell(r, h["Corte"]).value) == "PENDIENTE"
        ):
            contract = str(ws.cell(r, h["Contrato"]).value or "").strip()
            affected_contracts.add(contract)
            rows_to_delete.append(r)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    affected_contracts.update(
        manual_rows[x]["contrato"] for x in ids if x in manual_rows
    )
    for contract in affected_contracts:
        _write_visible_cell(
            wb,
            requirement,
            contract,
            manual_by_contract.get(contract, []),
        )
    return _save(wb)


def move_records_to_cut(data: bytes, requirement: str, record_ids: Iterable[str], cut_number: int) -> bytes:
    ids = {str(x) for x in record_ids}
    wb = _load(data)
    _ensure_technical_sheets(wb)
    falt = wb[FALTANTES_SHEET]
    fh = _headers(falt)

    manual_rows = {
        row["id"]: row
        for row in _manual_pending_rows_from_wb(wb, requirement)
        if row["id"] in ids
    }
    for manual_row in manual_rows.values():
        _append_manual_to_tech(wb, requirement, manual_row, int(cut_number))

    changed = len(manual_rows)
    for r in range(2, falt.max_row + 1):
        if (
            falt.cell(r, fh["Requerimiento"]).value == requirement
            and str(falt.cell(r, fh["ID"]).value) in ids
        ):
            falt.cell(r, fh["Corte"]).value = int(cut_number)
            changed += 1

    if changed == 0:
        raise ValueError("No se encontraron registros para agregar al corte.")

    _update_cut_count(wb, requirement, cut_number)
    return _save(wb)


def remove_records_from_cut(data: bytes, requirement: str, record_ids: Iterable[str], cut_number: int) -> bytes:
    ids = {str(x) for x in record_ids}
    wb = _load(data)
    _ensure_technical_sheets(wb)
    falt = wb[FALTANTES_SHEET]
    fh = _headers(falt)
    changed = 0
    for r in range(2, falt.max_row + 1):
        if falt.cell(r, fh["Requerimiento"]).value != requirement or str(falt.cell(r, fh["ID"]).value) not in ids:
            continue
        try:
            current_cut = int(falt.cell(r, fh["Corte"]).value)
        except (TypeError, ValueError):
            continue
        if current_cut == int(cut_number):
            falt.cell(r, fh["Corte"]).value = "PENDIENTE"
            changed += 1
    if changed == 0:
        raise ValueError("No se encontraron documentos para retirar del corte.")
    _update_cut_count(wb, requirement, cut_number)
    return _save(wb)


def delete_cut(data: bytes, requirement: str, cut_number: int) -> bytes:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    falt = wb[FALTANTES_SHEET]
    cuts = wb[CORTES_SHEET]
    fh = _headers(falt)
    ch = _headers(cuts)

    for r in range(2, falt.max_row + 1):
        if falt.cell(r, fh["Requerimiento"]).value != requirement:
            continue
        try:
            current_cut = int(falt.cell(r, fh["Corte"]).value)
        except (TypeError, ValueError):
            continue
        if current_cut == int(cut_number):
            falt.cell(r, fh["Corte"]).value = "PENDIENTE"

    deleted = False
    for r in range(2, cuts.max_row + 1):
        if cuts.cell(r, ch["Requerimiento"]).value == requirement and int(cuts.cell(r, ch["Corte"]).value) == int(cut_number):
            cuts.delete_rows(r, 1)
            deleted = True
            break
    if not deleted:
        raise ValueError("No se encontró el corte seleccionado.")
    return _save(wb)


def pending_summary_all(data: bytes) -> list[dict]:
    result: list[dict] = []
    for requirement in list_requirements(data):
        result.extend(pending_summary(data, requirement))
    return result


def _general_cut_marker(cut_number: int) -> str:
    return f"GENERAL:{int(cut_number)}"


def list_general_cuts(data: bytes) -> list[dict]:
    return list_cuts(data, "Todos")


def general_cut_details(data: bytes, cut_number: int) -> list[dict]:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    ws = wb[FALTANTES_SHEET]
    h = _headers(ws)
    marker = _general_cut_marker(cut_number)
    result = []

    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, h["Corte"]).value) != marker:
            continue
        documento = ws.cell(r, h["Documento"]).value
        especificacion = ws.cell(r, h["Especificacion"]).value if "Especificacion" in h else ""
        result.append({
            "id": ws.cell(r, h["ID"]).value,
            "requerimiento": ws.cell(r, h["Requerimiento"]).value,
            "contrato": ws.cell(r, h["Contrato"]).value,
            "auditor": ws.cell(r, h["Auditor"]).value,
            "codigo": ws.cell(r, h["Codigo_documento"]).value,
            "documento": documento,
            "especificacion": especificacion or "",
            "solicitud": _effective_document(documento, especificacion),
            "origen": _record_origin({
                "Origen": ws.cell(r, h["Origen"]).value if "Origen" in h else "",
                "Procedimiento": ws.cell(r, h["Procedimiento"]).value,
            }),
            "fecha": ws.cell(r, h["Fecha_deteccion"]).value,
            "corte": int(cut_number),
        })
    return result


def _update_general_cut_count(wb, cut_number: int) -> None:
    falt = wb[FALTANTES_SHEET]
    cuts = wb[CORTES_SHEET]
    fh = _headers(falt)
    ch = _headers(cuts)
    marker = _general_cut_marker(cut_number)
    count = 0

    for r in range(2, falt.max_row + 1):
        if str(falt.cell(r, fh["Corte"]).value) == marker:
            count += 1

    for r in range(2, cuts.max_row + 1):
        if (
            cuts.cell(r, ch["Requerimiento"]).value == "Todos"
            and int(cuts.cell(r, ch["Corte"]).value) == int(cut_number)
        ):
            cuts.cell(r, ch["Documentos"]).value = count
            break


def create_general_cut(data: bytes, user: str, cut_date) -> bytes:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    falt = wb[FALTANTES_SHEET]
    cuts = wb[CORTES_SHEET]
    fh = _headers(falt)
    ch = _headers(cuts)

    existing = [
        int(cuts.cell(r, ch["Corte"]).value)
        for r in range(2, cuts.max_row + 1)
        if cuts.cell(r, ch["Requerimiento"]).value == "Todos"
        and cuts.cell(r, ch["Corte"]).value
    ]
    next_cut = max(existing, default=0) + 1
    marker = _general_cut_marker(next_cut)

    manual_rows: list[tuple[str, dict]] = []
    for requirement in list_requirements(data):
        for row in _manual_pending_rows_from_wb(wb, requirement):
            manual_rows.append((requirement, row))

    for requirement, manual_row in manual_rows:
        _append_manual_to_tech(wb, requirement, manual_row, marker)

    affected = len(manual_rows)
    for r in range(2, falt.max_row + 1):
        if str(falt.cell(r, fh["Corte"]).value) == "PENDIENTE":
            falt.cell(r, fh["Corte"]).value = marker
            affected += 1

    if affected == 0:
        raise ValueError("No hay documentación pendiente para crear un corte.")

    row = [None] * cuts.max_column
    values = {
        "ID": uuid4().hex,
        "Requerimiento": "Todos",
        "Corte": next_cut,
        "Fecha": cut_date,
        "Creado_por": user,
        "Documentos": affected,
    }
    for name, value in values.items():
        row[ch[name] - 1] = value
    cuts.append(row)
    return _save(wb)


def move_records_to_general_cut(
    data: bytes,
    record_ids: Iterable[str],
    cut_number: int,
) -> bytes:
    ids = {str(x) for x in record_ids}
    wb = _load(data)
    _ensure_technical_sheets(wb)
    falt = wb[FALTANTES_SHEET]
    fh = _headers(falt)
    marker = _general_cut_marker(cut_number)

    manual_rows: dict[str, tuple[str, dict]] = {}
    for requirement in list_requirements(data):
        for row in _manual_pending_rows_from_wb(wb, requirement):
            if row["id"] in ids:
                manual_rows[row["id"]] = (requirement, row)

    for requirement, manual_row in manual_rows.values():
        _append_manual_to_tech(wb, requirement, manual_row, marker)

    changed = len(manual_rows)
    for r in range(2, falt.max_row + 1):
        if (
            str(falt.cell(r, fh["ID"]).value) in ids
            and str(falt.cell(r, fh["Corte"]).value) == "PENDIENTE"
        ):
            falt.cell(r, fh["Corte"]).value = marker
            changed += 1

    if changed == 0:
        raise ValueError("No se encontraron registros para agregar al corte.")

    _update_general_cut_count(wb, cut_number)
    return _save(wb)


def remove_records_from_general_cut(
    data: bytes,
    record_ids: Iterable[str],
    cut_number: int,
) -> bytes:
    ids = {str(x) for x in record_ids}
    wb = _load(data)
    _ensure_technical_sheets(wb)
    falt = wb[FALTANTES_SHEET]
    fh = _headers(falt)
    marker = _general_cut_marker(cut_number)
    changed = 0

    for r in range(2, falt.max_row + 1):
        if (
            str(falt.cell(r, fh["ID"]).value) in ids
            and str(falt.cell(r, fh["Corte"]).value) == marker
        ):
            falt.cell(r, fh["Corte"]).value = "PENDIENTE"
            changed += 1

    if changed == 0:
        raise ValueError("No se encontraron documentos para retirar del corte.")

    _update_general_cut_count(wb, cut_number)
    return _save(wb)


def delete_general_cut(data: bytes, cut_number: int) -> bytes:
    wb = _load(data)
    _ensure_technical_sheets(wb)
    falt = wb[FALTANTES_SHEET]
    cuts = wb[CORTES_SHEET]
    fh = _headers(falt)
    ch = _headers(cuts)
    marker = _general_cut_marker(cut_number)

    for r in range(2, falt.max_row + 1):
        if str(falt.cell(r, fh["Corte"]).value) == marker:
            falt.cell(r, fh["Corte"]).value = "PENDIENTE"

    deleted = False
    for r in range(2, cuts.max_row + 1):
        if (
            cuts.cell(r, ch["Requerimiento"]).value == "Todos"
            and int(cuts.cell(r, ch["Corte"]).value) == int(cut_number)
        ):
            cuts.delete_rows(r, 1)
            deleted = True
            break

    if not deleted:
        raise ValueError("No se encontró el corte seleccionado.")
    return _save(wb)


def documents_for_cuts(data: bytes, requirement: str, cuts_selected: Iterable[int]) -> list[dict]:
    selected = {int(x) for x in cuts_selected}
    result = []
    for row in get_faltantes(data, requirement):
        try:
            cut_int = int(row.get("Corte"))
        except (TypeError, ValueError):
            continue
        if cut_int in selected:
            result.append({
                "requerimiento": requirement,
                "contrato": row.get("Contrato"),
                "documento": row.get("Documento_efectivo"),
                "corte": cut_int,
            })
    return result
